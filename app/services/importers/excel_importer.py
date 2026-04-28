from datetime import datetime
import uuid

import openpyxl
from io import BytesIO
from decimal import Decimal


class ExcelRouteImporter:
    def __init__(self, file_contents):
        """
        file_contents: байты файла (например, из form.route_file.data.read())
        """
        # BytesIO позволяет openpyxl читать файл прямо из оперативной памяти
        self.wb = openpyxl.load_workbook(BytesIO(file_contents), data_only=True)
        self.sheet = self.wb.active
    
    def get_route_info(self):
        """Парсинг основной шапки маршрута (строки 1-7)"""
        try:
            # Читаем дату из ячейки (row=4, col=2)
            raw_start_date = self.sheet.cell(row=4, column=2).value
            
            # Если openpyxl уже распознал это как datetime, приводим к строке или ISO
            if isinstance(raw_start_date, datetime):
                start_date_str = raw_start_date.strftime("%Y-%m-%d")
            else:
                start_date_str = str(raw_start_date)
            
            # Используем метод cell(row, column) для точных координат
            data = {
                "carrier_name": self.sheet.cell(row=1, column=2).value,
                "carrier_id": str(self.sheet.cell(row=2, column=2).value).zfill(4),
                "unit_id": str(self.sheet.cell(row=3, column=2).value).zfill(4),
                "start_date": start_date_str,
                "route_name": self.sheet.cell(row=5, column=2).value,
                "route_number": self.sheet.cell(row=6, column=2).value,
                "transport_type": f"0x{self.sheet.cell(row=7, column=2).value}" 
                                if self.sheet.cell(row=7, column=2).value else "0x01",
                "region_code": str(66),
                "updated_at": datetime.now().isoformat() # Генерируем время импорта
            }
            return data
        except Exception as e:
            return {"error": f"Ошибка парсинга шапки: {str(e)}"}
    
    def get_full_data(self):
        route_info = self.get_route_info()
        tariff_blocks = []
        all_stops = []
        
        # Находим все строки-заголовки
        raw_blocks = []
        for row_idx in range(1, self.sheet.max_row + 1):
            cell_a = self.sheet.cell(row=row_idx, column=1).value
            if cell_a and "Набор серий" in str(cell_a):
                raw_blocks.append(row_idx)
        
        for i, start_row in enumerate(raw_blocks):
            # Генерируем UID, чтобы связать его с матрицей
            new_uid = f"t{uuid.uuid4().hex[:8]}"

            # Передаем только номер строки
            header = self._extract_tariff_header(start_row, i+1)
            if not header:
                continue
            
            header["uid"] = new_uid
            stops, matrix = self._parse_matrix_data(start_row)
            
            if i == 0:
                all_stops = stops
            
            header["matrix"] = matrix
            tariff_blocks.append(header)
            
        return {
            "route_info": route_info,
            "stops": all_stops,
            "tariffs": tariff_blocks
        }

    def _extract_tariff_header(self, row_idx, default_index):
        """
        Извлекает данные тарифа: 
        A: Ключевое слово (Набор серий)
        B: Техническая строка (1;02;21...)
        C: Название тарифа (Стоимость полного билета...)
        """
        raw_tech_str = str(self.sheet.cell(row=row_idx, column=2).value or "")
        # Читаем название напрямую из колонки C (3)
        tariff_name_val = self.sheet.cell(row=row_idx, column=3).value
        
        parts = [p.strip() for p in raw_tech_str.split(';') if p.strip()]
        
        if len(parts) < 2:
            return None

        table_type_code = parts[1]
        ss_codes_list = parts[2:] # Список кодов
        ss_codes_str = ";".join(ss_codes_list)

        # Если ячейка с названием пустая, ставим дефолт
        final_name = str(tariff_name_val).strip() if tariff_name_val else f"Тариф {default_index}"

        return {
            "tab_number": default_index,
            "tariff_name": final_name,
            "table_type_code": table_type_code,
            "ss_series_codes": ss_codes_str,
            "parsed_ss_codes_list": ss_codes_list,
            "start_row": row_idx
        }
    
    def _parse_matrix_data(self, start_row):
        stops = []
        prices_matrix = []
        current_row = start_row + 2
        
        while True:
            zone_idx_cell = self.sheet.cell(row=current_row, column=1).value
            if zone_idx_cell is None or not str(zone_idx_cell).isdigit():
                break
            
            stops.append(self.sheet.cell(row=current_row, column=2).value)
            
            row_prices = []
            # Читаем фиксированное количество колонок (по числу остановок)
            # Начинаем с 3-й колонки (C), где заголовок "0"
            for col_idx in range(3, 3 + 50): # 50 как предел
                price = self.sheet.cell(row=current_row, column=col_idx).value
                # Если заголовок столбца пустой — цены кончились
                header = self.sheet.cell(row=start_row + 1, column=col_idx).value
                if header is None:
                    break
                row_prices.append(price)
            
            prices_matrix.append(row_prices)
            current_row += 1
            
        return stops, prices_matrix
    
    def _detect_max_decimal_places(self, tariffs_data):
        """
        Анализирует все цены во всех матрицах и возвращает максимальное 
        количество знаков после запятой (от 0 до 2).
        """
        max_places = 0
        
        for tariff in tariffs_data:
            for row in tariff["matrix"]:
                for price in row:
                    if price is None or price == 0:
                        continue
                    
                    # Преобразуем в строку и смотрим, есть ли точка
                    str_price = str(price)
                    if "." in str_price:
                        # Считаем количество цифр после точки
                        places = len(str_price.split(".")[1])
                        if places > max_places:
                            max_places = places
                            
        # Ограничиваем результат бизнес-логикой (0, 1 или 2)
        return str(min(max_places, 2))
    
    def build_final_matrix(self, parsed_data):
        stops_count = len(parsed_data["stops"])
        
        # Создаем пустую матрицу
        final_matrix = [[{} for _ in range(stops_count)] for _ in range(stops_count)]
        
        for tariff in parsed_data["tariffs"]:
            tab_uid = tariff["uid"]
            raw_matrix = tariff["matrix"]
            
            for row_idx, row_prices in enumerate(raw_matrix):
                for col_idx, price in enumerate(row_prices):
                    if price is None or price == "":
                        continue
                        
                    try:
                        val = float(price)
                        
                        # ЛОГИКА ТРАНСПОРНИРОВАНИЯ:
                        # В Excel TRFZ цена "Из А в Б" часто лежит в нижнем углу.
                        # Чтобы она попала в верхний угол сайта (от меньшего индекса к большему),
                        # мы берем координаты так, чтобы всегда записывать в [min][max].
                        
                        target_row = min(row_idx, col_idx)
                        target_col = max(row_idx, col_idx)
                        
                        if target_row < stops_count and target_col < stops_count:
                            final_matrix[target_row][target_col][tab_uid] = val
                            
                    except (ValueError, TypeError):
                        continue
                            
        return final_matrix
    
    def get_formatted_route_data(self):
        """Возвращает данные в формате, полностью совместимом с общим роутом импорта"""
        raw = self.get_full_data()
        detected_places = self._detect_max_decimal_places(raw["tariffs"])
        
        # Подготавливаем остановки
        formatted_stops = [
            {"name": str(name).strip(), "km": "{:.2f}".format(float(i))} 
            for i, name in enumerate(raw["stops"], start=0)
        ]
        
        # Извлекаем исходные данные из Excel
        info = raw["route_info"]
        
        # СТРУКТУРА, КАК В TRFZ
        return {
            "common": {
                "region_code": str(info.get("region_code", "")),
                "carrier_id": str(info.get("carrier_id", "")),
                "unit_id": str(info.get("unit_id", "")),
                "decimal_places": str(detected_places),
            },
            "route_info": {
                "route_number": str(info.get("route_number", "")),
                "route_name": str(info.get("route_name", "")),
                "transport_type": str(info.get("transport_type", "0x01")),
                "start_date": info.get("start_date"),
                "updated_at": info.get("updated_at"),
            },
            "stops": formatted_stops,
            "tariff_tables": [
                {
                    "uid": t["uid"],
                    "tab_number": t["tab_number"],
                    "tariff_name": t["tariff_name"],
                    "table_type_code": t["table_type_code"],
                    "ss_series_codes": t["ss_series_codes"],
                    "parsed_ss_codes_list": t["parsed_ss_codes_list"],
                } for t in raw["tariffs"]
            ],
            "price_matrix": self.build_final_matrix(raw)
        }
